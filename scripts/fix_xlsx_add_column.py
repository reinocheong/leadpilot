"""Add Listing Type column to existing xlsx + re-classify sale posts"""
import openpyxl
from openpyxl.styles import Font, PatternFill
import re

xlsx = "/home/user/fb_data/fb_rentals.xlsx"
wb = openpyxl.load_workbook(xlsx)
ws = wb.active

# Check if Listing Type column already exists
headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
print(f"Current headers ({ws.max_column} cols): {headers}")

if "Listing Type" in headers:
    print("Column already exists, skipping insert.")
else:
    # Insert column C (after Property Name, before Property Type)
    ws.insert_cols(3)
    ws.cell(row=1, column=3, value="Listing Type")
    ws.cell(row=1, column=3).font = Font(bold=True)
    print("Inserted 'Listing Type' at column C")
    
    # Classify existing rows
    sale_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    for row in range(2, ws.max_row + 1):
        post_text = str(ws.cell(row, 13).value or "")  # Post Text = col 13 (was 12, now 13)
        listing_type = "出租"
        
        # Sale indicators
        sale_patterns = [
            r'\bfor\s*sale\b', r'\bFor Sale\b', r'\bFOR SALE\b',
            r'出售', r'售卖', r'卖屋', r'屋子出售',
            r'\bBrand New\b', r'\bbrand new\b',
        ]
        for pat in sale_patterns:
            if re.search(pat, post_text, re.IGNORECASE):
                listing_type = "出售"
                break
        
        # Price > 50000 is probably sale not rent
        rent_val = ws.cell(row, 7).value  # Rent was col 6, now col 7
        try:
            if int(rent_val) > 50000:
                listing_type = "出售"
        except:
            pass
        
        ws.cell(row, 3, value=listing_type)
        
        # Highlight sale rows
        if listing_type == "出售":
            ws.cell(row, 3).fill = sale_fill
    
    # Count
    sales = sum(1 for r in range(2, ws.max_row + 1) if ws.cell(r, 3).value == "出售")
    rents = ws.max_row - 1 - sales
    print(f"Classified: {rents} 出租 / {sales} 出售")

# Update headers reference
wb.save(xlsx)
print(f"Saved: {xlsx}")
print(f"Total rows: {ws.max_row}")

# Verify
wb2 = openpyxl.load_workbook(xlsx)
ws2 = wb2.active
print("\n=== Sample rows ===")
for r in [2, 3, 10, 20, 30, 40, 50]:
    if r <= ws2.max_row:
        lt = ws2.cell(r, 3).value or "-"
        pn = str(ws2.cell(r, 4).value or "")[:30]  # Property Name now col 4
        rent = ws2.cell(r, 7).value or "-"  # Rent now col 7
        link = str(ws2.cell(r, 9).value or "")[:60]  # Link now col 9
        print(f"Row {r}: [{lt}] {pn} | RM{rent} | {link}")
