import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

path = "/home/user/khl_content_calendar.xlsx"
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "内容日历"

header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
selected_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
wrap_align = Alignment(wrap_text=True, vertical='top')

headers = ["#", "周次", "状态", "建议日期", "题材", "标题方向", "内容角度", "主打产品", "风格", "完整帖文", "备注"]
col_widths = [4, 10, 10, 14, 14, 30, 40, 18, 10, 60, 20]

for col, (h, w) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border
    ws.column_dimensions[get_column_letter(col)].width = w

# Row 2: Mother's Day SELECTED — no 炸雪糕
row = [
    1, "May W2 (5/9-15)", "✅ 选题确认", "5月10日（周日）",
    "母亲节专题", "妈妈休息一天，用 KHL 和牛犒赏她",
    "软性情感——母亲节不煮饭，介绍和牛的便捷烹饪。Halal和牛品质感，档口/超市可转发",
    "日本和牛（A5宫崎和牛）", "生活化", "", "用户改：不要炸雪糕，只要和牛"
]
for col, val in enumerate(row, 1):
    cell = ws.cell(row=2, column=col, value=val)
    cell.alignment = wrap_align
    cell.border = thin_border
    cell.fill = selected_fill

# Rows 3-6: Other topics
topics = [
    [2, "May W2 (5/9-15)", "⏳ 待选", "5月12日（周二）", "和牛知识科普",
     "日本和牛等级怎么分？A5 vs A4 差别在哪？",
     "B2B专业硬核——帮餐厅老板选品，建立和牛进口商权威", "日本和牛", "专业", "", ""],
    [3, "May W2 (5/9-15)", "⏳ 待选", "5月13日（周三）", "品牌差异化故事",
     "从农场到冻库——KHL 玉米粒的完整旅程",
     "展示自种→加工→冷冻的全链条优势，区别于中间商", "冷冻玉米粒", "混合", "", ""],
    [4, "May W2 (5/9-15)", "⏳ 待选", "5月14日（周四）", "火锅料备货清单",
     "火锅料 Checklist——你的餐厅准备好了吗？",
     "火锅必备食材清单，自然带出产品线给餐厅老板参考", "冷冻肉类海鲜火锅料", "专业", "", ""],
    [5, "May W2 (5/9-15)", "⏳ 待选", "5月15日（周五）", "炸雪糕商业价值",
     "炸雪糕——你菜单上被忽略的高毛利甜品",
     "面向档口/餐厅，分析炸雪糕利润空间+出餐效率", "炸雪糕", "生活化+商业", "", ""],
]

for i, t in enumerate(topics, 3):
    for col, val in enumerate(t, 1):
        cell = ws.cell(row=i, column=col, value=val)
        cell.alignment = wrap_align
        cell.border = thin_border

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:K{2+len(topics)}"
wb.save(path)
print(f"OK: {path}")
