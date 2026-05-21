import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

path = "/home/user/khl_content_calendar.xlsx"
wb = openpyxl.load_workbook(path)
ws = wb.active

wrap_align = Alignment(wrap_text=True, vertical='top')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
drafted_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

# The full Mother's Day post (version F)
full_post = """妈妈的手，一辈子在忙。

小时候忙我们的三餐，
长大了忙我们的孩子。
厨房是她的战场，也是她表达爱的地方。

这个母亲节，让她从厨房退下来一次。
换我们为她做点什么——
哪怕只是一片煎得刚刚好的和牛。

🥩 KHL 日本 A5 宫崎和牛
海盐+黑胡椒，3 分钟上桌
不用厨艺，用心就好。

🌸 Halal 认证 · 批发供应 · 餐厅/超市/档口
📲 WhatsApp 私询

#KHLFOOD #母亲节 #谢谢妈妈"""

# Image description
image_desc = """配图方案A（推荐）：一双50岁左右妈妈的手，木砧板上放A5和牛（大理石油花清晰），暖黄厨房灯光，柔焦，海盐黑胡椒在旁边。情绪：warm, intimate, golden hour, no face.

配图方案C（备选）：两双筷子同时夹一片和牛——年轻手+年老手，人脸在画面外，暖色调日式美学。情绪：generations, sharing, Japanese aesthetic."""

# Update row 2
# Col 10 = 完整帖文 (J), Col 11 = 备注 (K)
ws.cell(row=2, column=10).value = full_post
ws.cell(row=2, column=10).alignment = wrap_align
ws.cell(row=2, column=10).border = thin_border

ws.cell(row=2, column=11).value = "用户选🅕风格 · " + image_desc
ws.cell(row=2, column=11).alignment = wrap_align
ws.cell(row=2, column=11).border = thin_border

# Update status
ws.cell(row=2, column=3).value = "✅ 帖文完成"
ws.cell(row=2, column=3).fill = drafted_fill

# Set row height for readability
ws.row_dimensions[2].height = 200

wb.save(path)
print("✅ xlsx updated — row 2: 帖文+配图 已写入")
print(f"   Path: {path}")
