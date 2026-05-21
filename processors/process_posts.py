import json
import openpyxl
from openpyxl.utils import get_column_letter
import re

# ── Load raw data ──
with open('/home/user/fb_data/fb_posts_raw.json') as f:
    posts = json.load(f)

# ── Load existing xlsx ──
try:
    wb = openpyxl.load_workbook('/home/user/fb_data/fb_rentals.xlsx')
    ws = wb.active
except FileNotFoundError:
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = ['Agent Name','Property Name','Property Type','Rooms','Furnishing',
               'Rent (RM)','Phone','Link','Photos','Remark','Scraped At','Post Text']
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h)

def extract_post_id(link):
    m = re.search(r'/posts/(\d+)/', link)
    return m.group(1) if m else link

# Build set of existing post IDs
existing_ids = set()
for r in range(2, ws.max_row + 1):
    link = str(ws.cell(r, 8).value or '')
    pid = extract_post_id(link)
    if pid:
        existing_ids.add(pid)

print(f"Existing post IDs in xlsx: {len(existing_ids)}")
print(f"Total raw posts: {len(posts)}")

# ── Extract structured data for each new post ──
extracted = []

# Post 10 (index 9): Xiiao Xiin - Studio near Tuas
extracted.append({
    'agent_name': 'Xiiao Xiin',
    'property_name': '',
    'property_type': 'Studio',
    'rooms': 'Studio',
    'furnishing': '全家私',
    'rent': 1900,
    'phone': '01128631058',
    'link': 'https://www.facebook.com/groups/1467428250213843/posts/4135858213370820/',
    'photos': '',
    'remark': 'Level 30, 靠近第二通道(Tuas), 6月可入住, 合约最少1年, 可商量',
    'scraped_at': '2026-05-09T08:52:14.615Z',
    'post_text': posts[9]['text']
})

# Post 11 (index 10): PeacefulZebra - Seeking 2 rooms with bathroom
extracted.append({
    'agent_name': 'PeacefulZebra',
    'property_name': '',
    'property_type': '',
    'rooms': '2间房带厕所',
    'furnishing': '',
    'rent': '',
    'phone': '',
    'link': 'https://www.facebook.com/groups/1467428250213843/posts/4134927096797265/',
    'photos': '',
    'remark': '[求租/Seeking] 找2间房带厕所。回复提及: Trellis 2bedroom, R&F 两房两卫中房出租',
    'scraped_at': '2026-05-09T08:52:14.616Z',
    'post_text': posts[10]['text']
})

# Post 12 (index 11): Teo ZhenYu - Paragon Suites Room
extracted.append({
    'agent_name': 'Teo ZhenYu',
    'property_name': 'Paragon Suites',
    'property_type': '房间',
    'rooms': 'Single Room',
    'furnishing': '全家私',
    'rent': 900,
    'phone': '',
    'link': 'https://www.facebook.com/groups/1467428250213843/posts/4135854110037897/',
    'photos': '',
    'remark': 'Fully renovated, Utilities & WiFi included (air-cond separate meter), Walking distance to CIQ, 5 min by car/motorbike, Direct rental from owner, Single male only, Ready to move in',
    'scraped_at': '2026-05-09T08:52:14.616Z',
    'post_text': posts[11]['text']
})

# Post 13 (index 12): Caren Chin - Sky Suites @ Bukit Meldrum
extracted.append({
    'agent_name': 'Caren Chin (REN47278)',
    'property_name': 'Sky Suites @ Bukit Meldrum',
    'property_type': '公寓',
    'rooms': '1房1厕',
    'furnishing': '全家私',
    'rent': 2000,
    'phone': '0127928873',
    'link': 'https://www.facebook.com/groups/1313487628797877/posts/4409628439183765/',
    'photos': '',
    'remark': 'New painting, New Furniture (no TV), 570 sqft, High Floor, 不到5分钟到Woodland关卡, Dep 2+1+0.5, Under touch up, 坡底公寓',
    'scraped_at': '2026-05-09T08:52:34.157Z',
    'post_text': posts[12]['text']
})

# Post 14 (index 13): Kenny Yap - Taman Universiti Factory
extracted.append({
    'agent_name': 'Kenny Yap (REN71756)',
    'property_name': 'Taman Universiti',
    'property_type': '排厂',
    'rooms': '',
    'furnishing': '',
    'rent': 4500,
    'phone': '0127676396',
    'link': 'https://www.facebook.com/groups/1313487628797877/posts/4409810695832206/',
    'photos': '',
    'remark': '1.5 Storey Terrace Factory, 面向大路旁, Jalan Perdagangan 1, Land size 25x80, Nego, 大学城一层半排厂',
    'scraped_at': '2026-05-09T08:52:34.158Z',
    'post_text': posts[13]['text']
})

# Post 15 (index 14): Ker Xin - Adda Height Cluster House
extracted.append({
    'agent_name': 'Ker Xin',
    'property_name': 'Adda Height (Maple Red)',
    'property_type': '排屋',
    'rooms': '4房4厕',
    'furnishing': '半家私',
    'rent': 2300,
    'phone': '',
    'link': 'https://www.facebook.com/groups/1313487628797877/posts/4409626375850638/',
    'photos': '',
    'remark': 'Double Storey Cluster House, 34x70, Dato Onn Unblock View, 2 aircond, Auto Gate, Kitchen Cabinet',
    'scraped_at': '2026-05-09T08:52:34.158Z',
    'post_text': posts[14]['text']
})

# ── Filter out already-existing posts ──
new_posts = []
for e in extracted:
    pid = extract_post_id(e['link'])
    if pid not in existing_ids:
        new_posts.append(e)
    else:
        print(f"SKIP (duplicate): {e['property_name'] or e['agent_name']} - {pid}")

print(f"New posts to add: {len(new_posts)}")

# ── Append new rows ──
next_row = ws.max_row + 1
for e in new_posts:
    ws.cell(next_row, 1, e['agent_name'])
    ws.cell(next_row, 2, e['property_name'])
    ws.cell(next_row, 3, e['property_type'])
    ws.cell(next_row, 4, e['rooms'])
    ws.cell(next_row, 5, e['furnishing'])
    ws.cell(next_row, 6, e['rent'] if e['rent'] != '' else '')
    ws.cell(next_row, 7, e['phone'])
    # Hyperlink for column H
    cell = ws.cell(next_row, 8)
    cell.value = e['link']
    cell.hyperlink = e['link']
    cell.style = 'Hyperlink'
    ws.cell(next_row, 9, e['photos'])
    ws.cell(next_row, 10, e['remark'])
    ws.cell(next_row, 11, e['scraped_at'])
    ws.cell(next_row, 12, e['post_text'])
    next_row += 1

# ── Make existing Link column hyperlinks too ──
for r in range(2, ws.max_row + 1):
    cell = ws.cell(r, 8)
    if cell.value and not cell.hyperlink:
        cell.hyperlink = str(cell.value)
        cell.style = 'Hyperlink'

# ── Formatting: auto-width ──
for col in range(1, 13):
    max_len = 0
    for r in range(1, ws.max_row + 1):
        val = ws.cell(r, col).value
        if val:
            max_len = max(max_len, min(len(str(val)), 60))
    ws.column_dimensions[get_column_letter(col)].width = max(max_len + 2, 12)

# ── Save ──
wb.save('/home/user/fb_data/fb_rentals.xlsx')
print(f"\nSaved! Total rows (incl header): {ws.max_row}")
print(f"Data rows: {ws.max_row - 1}")
print(f"New this run: {len(new_posts)}")
