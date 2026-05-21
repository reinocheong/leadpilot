#!/usr/bin/env python3
"""Process FB rental posts: extract structured fields, dedup, update xlsx."""

import json, re, os
from datetime import datetime

# ── Load raw posts ──────────────────────────────────────────────
with open('/home/user/fb_data/fb_posts_raw.json') as f:
    posts = json.load(f)

# ── Load existing xlsx ──────────────────────────────────────────
import openpyxl
from openpyxl.utils import get_column_letter

XLSX_PATH = '/home/user/fb_data/fb_rentals.xlsx'
HEADERS = [
    'Agent Name', 'Property Name', 'Property Type', 'Rooms',
    'Furnishing', 'Rent (RM)', 'Phone', 'Link', 'Photos',
    'Remark', 'Scraped At', 'Post Text'
]

def clean_link(link):
    """Remove tracking params but keep comment_id."""
    if not link:
        return ''
    link = re.sub(r'[?&]__cft__\[0\]=[^&]*', '', link)
    link = re.sub(r'[?&]__tn__=[^&]*', '', link)
    link = re.sub(r'[?&]$', '', link)
    return link

# Load existing workbook or create new
if os.path.exists(XLSX_PATH):
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb.active
    existing_links = set()
    for row in range(2, ws.max_row + 1):
        link = ws.cell(row, 8).value
        if link:
            existing_links.add(clean_link(link))
else:
    wb = openpyxl.Workbook()
    ws = wb.active
    for col_idx, header in enumerate(HEADERS, 1):
        ws.cell(1, col_idx, header)
    # Style header
    from openpyxl.styles import Font, PatternFill
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    for col_idx in range(1, len(HEADERS) + 1):
        ws.cell(1, col_idx).font = header_font
        ws.cell(1, col_idx).fill = header_fill
    existing_links = set()

print(f"Existing rows: {ws.max_row - 1}")
print(f"Existing unique links: {len(existing_links)}")

# ── Structured extraction (AI-extracted fields) ─────────────────
# Map: json_index -> {field: value}
# Only for posts that are actual rental listings or 求租 (not for-sale, not commercial)

EXTRACTED = {
    # idx 3: EmpatheticCranberry - SKS Pavilion (comment) — SKIP: duplicate of row 5 content
    
    # idx 5: WST Property Group - Sky Habitat/RNF/Twin Tower (comment) — SKIP: duplicate of row 6
    
    15: {  # Cheksiang Wee - D Inspire Residence
        'agent_name': 'Cheksiang Wee',
        'property_name': 'D Inspire Residence',
        'property_type': '小房',
        'rooms': '1间小房（共用厕所）',
        'furnishing': '',
        'rent': '',
        'phone': '',
        'remark': '位置超方便, 新马通勤友好, 5月1号可入住, 共用厕所/厨房/冰箱/洗衣机',
    },
    16: {  # 黄雄辉 - Johor Jaya 房间
        'agent_name': '黄雄辉',
        'property_name': 'Johor Jaya',
        'property_type': '房间',
        'rooms': '1间房间',
        'furnishing': '包冷气/热水器/双人床/衣橱',
        'rent': '',
        'phone': '',
        'remark': '包水电费, 包热水器, 冷气, 双人床, 衣橱, 屋主自租, 仅限女孩子',
    },
    17: {  # Lau John - Wanted Danga Bay/RNF
        'agent_name': 'Lau John',
        'property_name': '',
        'property_type': '',
        'rooms': '',
        'furnishing': '',
        'rent': '',
        'phone': '',
        'remark': '🔍求租: 找Danga Bay或RNF公寓 Studio, 屋主自租, 不要中介. 回复: SageLemur7249 2房1厕 no agent fee; MysticalSeahorse R&F中房有私人厕所对屋主',
    },
    18: {  # Jordan Kau - JP Perdana Cube 166
        'agent_name': 'Jordan Kau',
        'property_name': 'JP Perdana @ Cube 166',
        'property_type': '双层排屋',
        'rooms': '3房4厕',
        'furnishing': '半家私',
        'rent': 2400,
        'phone': '',
        'remark': '18x65 sqft, Dry & Wet Kitchens, Built-in Wardrobe, 2 Aircon, Shower Heater, Fridge, Sofa, Auto Gate, Unblocked View Facing Garden, Only Chinese Tenant, Jalan Jaya Putra 3, Skudai',
    },
    # idx 19: Factory — SKIP (commercial)
    # idx 20: Dato Onn For Sale — SKIP
    
    21: {  # Aidan Phang - Wanted 4房
        'agent_name': 'Aidan Phang',
        'property_name': '',
        'property_type': '',
        'rooms': '',
        'furnishing': '',
        'rent': '',
        'phone': '',
        'remark': '🔍求租: 想租整间, 4房, budget 2000+, 4人男生, 在新加坡工作, 希望有冷气. 回复: 苏薇云 0164142768 4房4厕; SageLemur One49 3BR 2k partial furnish',
    },
    22: {  # Kai Ze - Wanted Studio
        'agent_name': 'Kai Ze',
        'property_name': '',
        'property_type': '',
        'rooms': '',
        'furnishing': '',
        'rent': '',
        'phone': '',
        'remark': '🔍求租: 找Studio/整间, Bukit Indah/Perling/Iskandar/Forest City, Budget RM1500以内, 下个月中或尾入住. 回复: Ivan Lo Sky Trees Bukit Indah fully furnished 包水电wifi',
    },
    23: {  # Lau John - Wanted Studio CIQ
        'agent_name': 'Lau John',
        'property_name': '',
        'property_type': '',
        'rooms': '',
        'furnishing': '',
        'rent': '',
        'phone': '',
        'remark': '🔍求租: 找Studio 2人住, budget 1400-1600, 近CIQ. 回复: CK Lim 有房; Kc Khoo SKS Pavilion Studio 全家私 RM2100 包车位 0179711541',
    },
    24: {  # Kc Khoo - SKS Pavilion Studio
        'agent_name': 'Kc Khoo',
        'property_name': 'SKS Pavilion',
        'property_type': 'Studio',
        'rooms': 'Studio',
        'furnishing': '全家私',
        'rent': 2100,
        'phone': '0179711541',
        'remark': '走路就能到CIQ关卡, 包车位, 租约至少一年以上',
    },
    25: {  # 匿名互动者 - Wanted Bukit Indah
        'agent_name': '匿名互动者',
        'property_name': '',
        'property_type': '',
        'rooms': '',
        'furnishing': '',
        'rent': '',
        'phone': '',
        'remark': '🔍求租: Bukit Indah/Skudai区, 情侣入住爱干净, 找有独立厕所, 包水电, Budget RM750以下, 不长期回来. 回复: Soo Mei Wong 振林山; Wong Hieng Nusa Sentral房间',
    },
    26: {  # Mable Ng - Permas 房间
        'agent_name': 'Mable Ng',
        'property_name': 'Permas (Permas 4)',
        'property_type': '房间',
        'rooms': '1间普通房',
        'furnishing': '全家私',
        'rent': '',
        'phone': '',
        'remark': '双层排楼, Permas 4, 包水电冷气网, 全新床架/床/衣橱',
    },
    27: {  # Joyce JQ - Country Garden Danga Bay
        'agent_name': 'Joyce JQ',
        'property_name': 'Country Garden Danga Bay',
        'property_type': '公寓',
        'rooms': '多种选择',
        'furnishing': '全家私',
        'rent': 700,
        'phone': '01160615622',
        'remark': '5 mins to CIQ, LOW DEPO, No Agent Fees, Motor Parking & Basic Utility included, Opposite Paragon Residences, Near Citywoods/Goldensands. Common Room RM700, Balcony Room RM800, Submaster RM1100, Master RM1300, Studio RM1800, 2BR RM2200, 3BR RM3000. Ready Move In.',
    },
    # idx 28: Austin Crest For Sale — SKIP
    # idx 29: Setia Indah For Sale — SKIP
}

# ── Process new posts ───────────────────────────────────────────
new_rows = 0
skipped_dup = 0
skipped_other = 0

for idx, post in enumerate(posts):
    link = clean_link(post.get('link', ''))
    
    # Check if link exists
    if link in existing_links:
        skipped_dup += 1
        continue
    
    # Check if we have extracted data for this index
    if idx not in EXTRACTED:
        skipped_other += 1
        continue
    
    data = EXTRACTED[idx]
    
    # Build row
    next_row = ws.max_row + 1
    row_data = [
        data.get('agent_name', post.get('agent_name', '')),
        data.get('property_name', ''),
        data.get('property_type', ''),
        data.get('rooms', ''),
        data.get('furnishing', ''),
        data.get('rent', ''),
        data.get('phone', post.get('phone', '')),
        link,
        ', '.join(post.get('photos', [])) if post.get('photos') else '',
        data.get('remark', ''),
        post.get('scraped_at', datetime.now().isoformat()),
        post.get('text', ''),
    ]
    
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(next_row, col_idx, value)
    
    # Make link column clickable
    link_cell = ws.cell(next_row, 8)
    if link:
        link_cell.hyperlink = link
        link_cell.font = openpyxl.styles.Font(color='0563C1', underline='single')
    
    existing_links.add(link)
    new_rows += 1

# ── Auto-fit column widths ──────────────────────────────────────
for col_idx in range(1, len(HEADERS) + 1):
    max_width = len(HEADERS[col_idx - 1])
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row, col_idx).value
        if val:
            max_width = max(max_width, min(len(str(val)), 60))
    ws.column_dimensions[get_column_letter(col_idx)].width = max_width + 2

# ── Save ────────────────────────────────────────────────────────
wb.save(XLSX_PATH)
print(f"\nSaved: {XLSX_PATH}")
print(f"  Total rows (excl. header): {ws.max_row - 1}")
print(f"  New rows added: {new_rows}")
print(f"  Duplicates skipped: {skipped_dup}")
print(f"  Other skipped (for-sale/commercial/no-data): {skipped_other}")
